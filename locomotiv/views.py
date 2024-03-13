import openpyxl as openpyxl
from rest_framework.response import Response
from rest_framework import generics, status

from locomotiv.models import Locomotiv, TotalDataVagon
from locomotiv.serializers import NumberSerializer, TotalDataSerializer, InputDataSerializer, UploadVagonNumberSerializer
from locomotiv.utils import get_vagon_data, is_true
from .usecases import UseCases


class VagonDataListView(generics.CreateAPIView):
    queryset = Locomotiv.objects.all()
    serializer_class = NumberSerializer

    def get(self, *args, **kwargs):
        queryset = TotalDataVagon.objects.all()
        serializer = TotalDataSerializer(queryset, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

    def post(self, *args, **kwargs):
        serializer = NumberSerializer(data=self.request.POST)
        serializer.is_valid(raise_exception=True)
        number = serializer.data['number']
        if number[0] == '0' or number[0] == '1':
            return Response({"error_message": "Vagon raqami 0 yoki 1 bilan boshlanmasligi kerak"},
                            status=status.HTTP_400_BAD_REQUEST)
        load_weight = serializer.data['load_weight']
        full_data = get_vagon_data(number, load_weight)

        input_data = {
            'number_of_arrow': full_data['number_of_arrow'],
            'netto_vagon': full_data['netto_vagon'],
            'length_vagon': full_data['length_vagon']
        }
        input_serializer = InputDataSerializer(data=input_data)
        input_serializer.is_valid(raise_exception=True)

        serializer = TotalDataSerializer(data=full_data)
        serializer.is_valid(raise_exception=True)
        serializer.save()

        return Response(full_data, status=status.HTTP_200_OK)


class UploadVagonNumberDataView(generics.GenericAPIView):
    serializer_class = UploadVagonNumberSerializer
    queryset = TotalDataVagon.objects.all()

    def post(self, request, *args, **kwargs):
        try:
            file = self.request.FILES.get('file', None)
        except:
            return Response({
                "message": "File is not provided",
                "code": "bad_request"
            }, status=status.HTTP_400_BAD_REQUEST)
        if not file:
            return Response({
                "message": "File not found",
                "code": "not_found"
            }, status=status.HTTP_404_NOT_FOUND)
        dataframe = openpyxl.load_workbook(file)

        # Define variable to read sheet
        dataframe1 = dataframe.active

        # Iterate the loop to read the cell values
        for row in range(0, dataframe1.max_row):
            number = 0
            weight = 0
            i = 0
            if row == 0:
                continue
            for col in dataframe1.iter_cols(1, 2):
                if i == 0:
                    number = col[row].value
                else:
                    weight = col[row].value
                i += 1
                print(col[row].value)

            full_data = get_vagon_data(number, weight)
            serializer = TotalDataSerializer(data=full_data)
            serializer.is_valid(raise_exception=True)
            serializer.save()

        return Response('ok')


class NewAppCalculatingResistanceAPIVIew(generics.ListAPIView):
    queryset = TotalDataVagon.objects.all()

    def get(self, request, *args, **kwargs):
        """
        bu method, NEW APP uchun vagonlarni
        guruhlash, shu guruhlar bo'yicha qarshiliklarni
        aniqlash uchun ishlatiladi
        """
        articulated_road = is_true(self.request.query_params.get('articulated_road', True))
        vagons = self.get_queryset()
        vagons_group_data = UseCases.making_vagon_groups(vagons)
        data = []
        for capacity in range(1, 81):
            resistances_in_capacity = UseCases.calculate_resistance_for_all_groups(
                capacity, vagons_group_data, articulated_road
            )
            data.append(
                resistances_in_capacity
            )
        return Response({
            "vagons_group_data": vagons_group_data,
            "resistances": data
        })


